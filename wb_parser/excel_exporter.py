from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    @staticmethod
    def export_to_excel(products, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"
        
        headers = [
            'Ссылка на товар', 'Артикул', 'Название', 'Цена', 'Описание',
            'Ссылки на изображения', 'Характеристики', 'Название продавца',
            'Ссылка на продавца', 'Размеры', 'Остатки', 
            'Рейтинг', 'Количество отзывов'
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", 
                                  fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_num, product in enumerate(products, 2):
            characteristics_str = '\n'.join([f'{k}: {v}' 
                                for k, v in product.characteristics.items()])
            
            ws.cell(row=row_num, column=1, value=product.url)
            ws.cell(row=row_num, column=2, value=product.article)
            ws.cell(row=row_num, column=3, value=product.name)
            ws.cell(row=row_num, column=4, value=product.price)
            ws.cell(row=row_num, column=5, value=product.description)
            ws.cell(row=row_num, column=6, value=product.image_urls)
            ws.cell(row=row_num, column=7, value=characteristics_str)
            ws.cell(row=row_num, column=8, value=product.seller_name)
            ws.cell(row=row_num, column=9, value=product.seller_url)
            ws.cell(row=row_num, column=10, value=product.sizes)
            ws.cell(row=row_num, column=11, value=product.stock)
            ws.cell(row=row_num, column=12, value=product.rating)
            ws.cell(row=row_num, column=13, value=product.reviews_count)
                
        for row in ws.iter_rows(min_row=2, max_row=len(products) + 1, 
                                min_col=5, max_col=7):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        wb.save(filename)
        print(f"Файл сохранен: {filename}")

    @staticmethod
    def auto_adjust_column_widths(worksheet, min_width=10, max_width=50):

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    if isinstance(cell.value, str):
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            cell_length = len(line)
                            if cell.font and cell.font.bold:
                                cell_length = int(cell_length * 1.1)
                            max_length = max(max_length, cell_length)
                    else:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
            
            adjusted_width = min(max_length + 2, max_width)
            adjusted_width = max(adjusted_width, min_width)
            
            worksheet.column_dimensions[column_letter].width = adjusted_width